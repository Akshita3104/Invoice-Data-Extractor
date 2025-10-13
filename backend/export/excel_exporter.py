"""
Excel Exporter
Exports invoice data to Excel format with formatting and validation reports
Enhanced version with better formatting and multiple sheets
"""

import os
import pandas as pd
from typing import Dict, List, Optional
from datetime import datetime


class ExcelExporter:
    """
    Exports invoice data to Excel format
    """
    
    def __init__(self):
        """Initialize Excel exporter"""
        self.default_columns = [
            'Goods Description',
            'HSN/SAC Code',
            'Quantity',
            'Weight',
            'Rate',
            'Amount',
            'Company Name',
            'Invoice Number',
            'FSSAI Number',
            'Date of Invoice'
        ]
    
    def export(
        self,
        data: List[Dict],
        output_folder: str,
        filename: str = "invoice_data.xlsx",
        include_validation: bool = True,
        validation_issues: Dict = None,
        confidence_scores: Dict = None
    ) -> str:
        """
        Export data to Excel
        
        Args:
            data: List of invoice items
            output_folder: Output folder path
            filename: Output filename
            include_validation: Include validation sheet
            validation_issues: Validation issues from validators
            confidence_scores: Confidence scores
            
        Returns:
            Path to exported file
        """
        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = os.path.join(output_folder, filename)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main data sheet
            self._write_data_sheet(data, writer)
            
            # Validation sheet (if requested)
            if include_validation and validation_issues:
                self._write_validation_sheet(validation_issues, writer)
            
            # Confidence sheet (if available)
            if confidence_scores:
                self._write_confidence_sheet(confidence_scores, writer)
            
            # Summary sheet
            self._write_summary_sheet(data, validation_issues, confidence_scores, writer)
        
        # Apply formatting
        self._apply_formatting(output_path)
        
        print(f"Excel file exported to: {output_path}")
        
        return output_path
    
    def _write_data_sheet(self, data: List[Dict], writer):
        """
        Write main data sheet
        """
        if not data:
            # Create empty DataFrame with headers
            df = pd.DataFrame(columns=self.default_columns)
        else:
            # Convert to DataFrame
            df = pd.DataFrame(data)
            
            # Ensure all default columns exist
            for col in self.default_columns:
                if col not in df.columns:
                    df[col] = 'N/A'
            
            # Reorder columns
            existing_cols = [col for col in self.default_columns if col in df.columns]
            other_cols = [col for col in df.columns if col not in self.default_columns]
            df = df[existing_cols + other_cols]
            
            # Convert weight to kg if needed
            if 'Weight' in df.columns:
                df['Weight'] = df['Weight'].apply(self._convert_weight_to_kg)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Invoice Data', index=False)
    
    def _write_validation_sheet(self, validation_issues: Dict, writer):
        """
        Write validation issues sheet
        """
        all_issues = []
        
        for validator_type, issues in validation_issues.items():
            for issue in issues:
                all_issues.append({
                    'Validator': validator_type.capitalize(),
                    'Type': issue.get('type', 'unknown'),
                    'Severity': issue.get('severity', 'unknown'),
                    'Message': issue.get('message', ''),
                    'Item': issue.get('item', 'N/A')
                })
        
        if all_issues:
            df = pd.DataFrame(all_issues)
            df.to_excel(writer, sheet_name='Validation Issues', index=False)
        else:
            # No issues - write success message
            df = pd.DataFrame({
                'Message': ['All validations passed! No issues detected.']
            })
            df.to_excel(writer, sheet_name='Validation Issues', index=False)
    
    def _write_confidence_sheet(self, confidence_scores: Dict, writer):
        """
        Write confidence scores sheet
        """
        # Overall metrics
        overall_data = {
            'Metric': [
                'Overall Confidence',
                'Quality Level',
                'Completeness',
                'Needs Review',
                'Ready for Processing'
            ],
            'Value': [
                f"{confidence_scores.get('overall_confidence', 0):.1%}",
                confidence_scores.get('quality_level', 'unknown').upper(),
                f"{confidence_scores.get('completeness', 0):.1%}",
                'YES' if confidence_scores.get('needs_review', False) else 'NO',
                'YES' if confidence_scores.get('ready_for_processing', False) else 'NO'
            ]
        }
        
        df_overall = pd.DataFrame(overall_data)
        df_overall.to_excel(writer, sheet_name='Confidence Scores', index=False, startrow=0)
        
        # Validation breakdown
        validation_scores = confidence_scores.get('validation_scores', {})
        if validation_scores:
            val_data = {
                'Validation Type': list(validation_scores.keys()),
                'Score': [f"{v:.1%}" for v in validation_scores.values()]
            }
            df_val = pd.DataFrame(val_data)
            df_val.to_excel(writer, sheet_name='Confidence Scores', index=False, startrow=len(overall_data['Metric']) + 3)
        
        # Field-level confidence
        field_conf = confidence_scores.get('field_confidence', {})
        if field_conf:
            field_data = {
                'Field': list(field_conf.keys()),
                'Confidence': [f"{v:.1%}" for v in field_conf.values()]
            }
            df_field = pd.DataFrame(field_data)
            start_row = len(overall_data['Metric']) + len(validation_scores) + 6
            df_field.to_excel(writer, sheet_name='Confidence Scores', index=False, startrow=start_row)
    
    def _write_summary_sheet(
        self,
        data: List[Dict],
        validation_issues: Dict = None,
        confidence_scores: Dict = None,
        writer = None
    ):
        """
        Write summary sheet
        """
        summary_data = []
        
        # Basic stats
        summary_data.append({'Metric': 'Total Items', 'Value': len(data)})
        
        # Calculate total amount
        total_amount = 0
        for item in data:
            amount = self._parse_number(item.get('Amount', ''))
            if amount:
                total_amount += amount
        
        summary_data.append({'Metric': 'Total Amount', 'Value': f"₹{total_amount:,.2f}"})
        
        # Unique invoice numbers
        invoice_numbers = set(item.get('Invoice Number', '') for item in data if item.get('Invoice Number') != 'N/A')
        summary_data.append({'Metric': 'Invoices Processed', 'Value': len(invoice_numbers)})
        
        # Unique companies
        companies = set(item.get('Company Name', '') for item in data if item.get('Company Name') != 'N/A')
        summary_data.append({'Metric': 'Unique Companies', 'Value': len(companies)})
        
        # Validation summary
        if validation_issues:
            total_issues = sum(len(issues) for issues in validation_issues.values())
            summary_data.append({'Metric': 'Validation Issues', 'Value': total_issues})
        
        # Confidence summary
        if confidence_scores:
            summary_data.append({
                'Metric': 'Overall Confidence',
                'Value': f"{confidence_scores.get('overall_confidence', 0):.1%}"
            })
            summary_data.append({
                'Metric': 'Quality Level',
                'Value': confidence_scores.get('quality_level', 'unknown').upper()
            })
        
        # Export timestamp
        summary_data.append({
            'Metric': 'Export Date',
            'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _apply_formatting(self, output_path: str):
        """
        Apply formatting to Excel file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Load workbook
            wb = load_workbook(output_path)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format header row
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save formatted workbook
            wb.save(output_path)
            
        except ImportError:
            print("Warning: openpyxl not available for formatting")
        except Exception as e:
            print(f"Warning: Could not apply formatting: {e}")
    
    def _convert_weight_to_kg(self, weight_str: str) -> str:
        """
        Convert weight to kg
        """
        if not weight_str or weight_str == 'N/A':
            return weight_str
        
        weight_str = str(weight_str).upper().replace(' ', '')
        
        # Extract number
        import re
        match = re.search(r'(\d+\.?\d*)', weight_str)
        if not match:
            return weight_str
        
        value = float(match.group())
        
        # Convert based on unit
        if 'QTL' in weight_str or 'QUINTAL' in weight_str:
            value = value * 100
        elif 'TON' in weight_str or 'MT' in weight_str:
            value = value * 1000
        elif 'G' in weight_str and 'KG' not in weight_str:
            value = value / 1000
        
        return f"{value} KG"
    
    def _parse_number(self, value: str) -> Optional[float]:
        """
        Parse number from string
        """
        if not value or value == 'N/A':
            return None
        
        import re
        
        # Remove currency and commas
        cleaned = re.sub(r'[₹$€£¥,\s]', '', str(value))
        cleaned = re.sub(r'Rs\.?', '', cleaned, flags=re.IGNORECASE)
        
        # Extract number
        match = re.search(r'-?\d+\.?\d*', cleaned)
        if match:
            try:
                return float(match.group())
            except ValueError:
                return None
        
        return None
    
    def export_simple(
        self,
        data: List[Dict],
        output_path: str
    ) -> str:
        """
        Simple export without validation (backward compatible)
        
        Args:
            data: List of invoice items
            output_path: Full output path including filename
            
        Returns:
            Path to exported file
        """
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        if not data:
            df = pd.DataFrame(columns=self.default_columns)
        else:
            df = pd.DataFrame(data)
            
            # Convert weight
            if 'Weight' in df.columns:
                df['Weight'] = df['Weight'].apply(self._convert_weight_to_kg)
        
        # Export
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        print(f"Excel file saved at {output_path}")
        
        return output_path